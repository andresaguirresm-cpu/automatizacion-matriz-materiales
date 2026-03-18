"""
excel_writer.py — Generación del Excel con formato, merges y colores por ETAPA.
"""
import io
from itertools import groupby
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from .matrix_builder import OutputRow


# Colores por índice de ETAPA (cíclico)
ETAPA_COLORS = {
    0: "BDD7EE",
    1: "DDEBF7",
    2: "FFF2CC",
    3: "E2EFDA",
}
HEADER_COLOR = "1F3864"      # Azul oscuro para el título
HEADER_FONT_COLOR = "FFFFFF"
COL_HEADER_COLOR = "2F5496"  # Azul medio para cabeceras de columnas
COL_HEADER_FONT = "FFFFFF"
SEPARATOR_COLOR = "D9D9D9"   # Gris claro para filas separadoras

COLUMNS = [
    ("ETAPA", 18),
    ("SUB ETAPA", 18),
    ("MEDIO", 16),
    ("FORMATO", 22),
    ("MENSAJE", 30),
    ("SPECCS", 35),
    ("PESO", 16),
    ("EXTENSIÓN", 14),
    ("TEXTO", 30),
    ("SEGUNDAJE IDEAL", 18),
]

# Letras de columna (A-J)
COL_LETTERS = [get_column_letter(i + 1) for i in range(len(COLUMNS))]


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border_thin() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _border_separator() -> Border:
    med = Side(style="medium", color="9E9E9E")
    return Border(top=med, bottom=med)


def write_excel(
    rows: list[OutputRow],
    campaign_name: str = "CAMPAÑA",
) -> bytes:
    """
    Genera el Excel formateado y retorna los bytes del archivo.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "MATRIZ DIGITAL"

    # ── Fila 1 vacía (opción estética)
    ws.row_dimensions[1].height = 6

    # ── Fila 2: Título principal (merged A:J)
    title_text = f"MATRIZ DIGITAL – {campaign_name.upper()}"
    ws.merge_cells("A2:J2")
    title_cell = ws["A2"]
    title_cell.value = title_text
    title_cell.fill = _fill(HEADER_COLOR)
    title_cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30

    # ── Fila 3: Headers de columnas
    header_row = 3
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = _fill(COL_HEADER_COLOR)
        cell.font = Font(bold=True, color=COL_HEADER_FONT, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[header_row].height = 22

    # ── Filas de datos agrupadas por ETAPA
    current_row = 4
    etapa_color_idx = 0

    # Agrupar por etapa manteniendo orden de aparición
    etapas_order = []
    seen = set()
    for r in rows:
        if r.etapa not in seen:
            etapas_order.append(r.etapa)
            seen.add(r.etapa)

    for etapa in etapas_order:
        etapa_rows = [r for r in rows if r.etapa == etapa]
        hex_color = ETAPA_COLORS[etapa_color_idx % len(ETAPA_COLORS)]
        fill = _fill(hex_color)

        etapa_start_row = current_row

        # Agrupar por sub_etapa dentro de la etapa
        sub_etapas_order = []
        seen_sub = set()
        for r in etapa_rows:
            if r.sub_etapa not in seen_sub:
                sub_etapas_order.append(r.sub_etapa)
                seen_sub.add(r.sub_etapa)

        for sub_etapa in sub_etapas_order:
            sub_rows = [r for r in etapa_rows if r.sub_etapa == sub_etapa]
            sub_start_row = current_row

            for row_data in sub_rows:
                # Escribir celdas de datos
                values = [
                    row_data.etapa,
                    row_data.sub_etapa,
                    row_data.medio,
                    row_data.formato,
                    row_data.mensaje,
                    row_data.speccs,
                    row_data.peso,
                    row_data.extension,
                    row_data.texto,
                    row_data.segundaje_ideal,
                ]
                for col_idx, val in enumerate(values, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.fill = fill
                    cell.font = Font(size=9)
                    cell.alignment = Alignment(
                        vertical="center",
                        wrap_text=True,
                        horizontal="center" if col_idx in (1, 2, 3, 7, 8, 9, 10) else "left",
                    )
                    cell.border = _border_thin()

                # Regla especial YOUTUBE: si PESO = "ENVIAR A MEDIOS EL LINK..."
                if (
                    "YOUTUBE" in row_data.medio.upper()
                    and "ENVIAR" in row_data.peso.upper()
                ):
                    ws.merge_cells(
                        start_row=current_row, start_column=7,
                        end_row=current_row, end_column=9
                    )

                ws.row_dimensions[current_row].height = 45
                current_row += 1

            # Merge SUB ETAPA (col B) para todo el sub-bloque
            if current_row - sub_start_row > 1:
                ws.merge_cells(
                    start_row=sub_start_row, start_column=2,
                    end_row=current_row - 1, end_column=2
                )
                sub_cell = ws.cell(row=sub_start_row, column=2)
                sub_cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            # Merge MENSAJE (col E) por grupos consecutivos del mismo mensaje
            # (no por todo el sub-bloque, para que mensajes distintos se vean separados)
            msg_group_start = sub_start_row
            msg_group_val = sub_rows[0].mensaje if sub_rows else ""
            for i, row_data in enumerate(sub_rows):
                excel_row = sub_start_row + i
                is_last = (i == len(sub_rows) - 1)
                next_msg = sub_rows[i + 1].mensaje if not is_last else None
                if is_last or next_msg != row_data.mensaje:
                    if excel_row > msg_group_start:
                        ws.merge_cells(
                            start_row=msg_group_start, start_column=5,
                            end_row=excel_row, end_column=5
                        )
                    msg_cell = ws.cell(row=msg_group_start, column=5)
                    msg_cell.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                    msg_group_start = excel_row + 1
                    msg_group_val = next_msg or ""

        # Merge ETAPA (col A) para todo el bloque de la etapa
        if current_row - etapa_start_row > 1:
            ws.merge_cells(
                start_row=etapa_start_row, start_column=1,
                end_row=current_row - 1, end_column=1
            )
            etapa_cell = ws.cell(row=etapa_start_row, column=1)
            etapa_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            etapa_cell.font = Font(bold=True, size=9)

        # Fila separadora entre ETAPAs
        for col_idx in range(1, len(COLUMNS) + 1):
            sep_cell = ws.cell(row=current_row, column=col_idx, value="")
            sep_cell.fill = _fill(SEPARATOR_COLOR)
            sep_cell.border = _border_thin()
        ws.row_dimensions[current_row].height = 4
        current_row += 1

        etapa_color_idx += 1

    # Freeze panes debajo de los headers
    ws.freeze_panes = "A4"

    # Serializar a bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
